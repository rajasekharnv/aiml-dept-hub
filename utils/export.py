import pandas as pd
import io
import unicodedata
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# UNICODE TO ASCII SANITIZATION HELPER
# =====================================================================

def sanitize_text(text: str) -> str:
    """
    Sanitizes Unicode special characters (em dashes, smart quotes, bullets, ellipses, etc.)
    by replacing them with their ASCII equivalents and normalization to avoid FPDF encoding exceptions.
    """
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
        
    # Common Unicode character mappings to ASCII equivalents
    replacements = {
        "\u2014": "-",   # em dash
        "\u2013": "-",   # en dash
        "\u201c": '"',   # smart left double quote
        "\u201d": '"',   # smart right double quote
        "\u2018": "'",   # smart left single quote
        "\u2019": "'",   # smart right single quote
        "\u2022": "*",   # bullet
        "\u2026": "...", # ellipsis
        "\u00a0": " ",   # non-breaking space
        "\u2212": "-",   # minus sign
        "\u2122": "(TM)",
        "\u00ae": "(R)",
        "\u00a9": "(C)",
        "\u2032": "'",   # prime
        "\u2033": '"',   # double prime
        "\u201b": "'",   # reversed single quote
        "\u201f": '"',   # reversed double quote
        "\u201a": "'",   # single low-9 quote
        "\u201e": '"',   # double low-9 quote
    }
    
    for uni_char, ascii_char in replacements.items():
        text = text.replace(uni_char, ascii_char)
        
    # Normalize to NFKD to decompose accented characters (e.g. é -> e)
    normalized = unicodedata.normalize('NFKD', text)
    # Encode to ASCII, ignoring any remaining unsupported characters, and decode back to string
    return normalized.encode('ascii', 'ignore').decode('ascii')


# =====================================================================
# GLOBAL MONKEYPATCH: SANITIZE TIMEZONES FOR EXCEL EXPORT
# =====================================================================
_original_to_excel = pd.DataFrame.to_excel

def _patched_to_excel(self, *args, **kwargs):
    df_clean = self.copy()
    
    if isinstance(df_clean.index, pd.DatetimeIndex) and df_clean.index.tz is not None:
        df_clean.index = df_clean.index.tz_localize(None)
        
    for col in df_clean.columns:
        if pd.api.types.is_datetime64_any_dtype(df_clean[col]):
            if hasattr(df_clean[col].dt, "tz") and df_clean[col].dt.tz is not None:
                try:
                    df_clean[col] = df_clean[col].dt.tz_localize(None)
                except Exception:
                    try:
                        df_clean[col] = df_clean[col].dt.tz_convert(None).dt.tz_localize(None)
                    except Exception:
                        pass
        else:
            def make_tz_unaware(val):
                if hasattr(val, "tzinfo") and val.tzinfo is not None:
                    try:
                        return val.replace(tzinfo=None)
                    except Exception:
                        return str(val)
                return val
            df_clean[col] = df_clean[col].apply(make_tz_unaware)
            
    return _original_to_excel(df_clean, *args, **kwargs)

pd.DataFrame.to_excel = _patched_to_excel

# =====================================================================
# BACKWARDS COMPATIBILITY EXPORTS
# =====================================================================

def export_to_excel(data: list[dict]) -> bytes:
    """
    Exports a list of dictionaries to an Excel file (bytes).
    Uses pandas and openpyxl internally.
    """
    if not data:
        data = [{"Message": "No data available to export"}]
        
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    # Write to Excel in memory
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
    return output.getvalue()

def export_to_pdf(title: str, content: list[str]) -> bytes:
    """
    Generates a PDF document from text paragraphs.
    Returns the file content as bytes.
    """
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font("Helvetica", style="B", size=16)
    pdf.cell(0, 10, sanitize_text(title), new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(10)
    
    # Body text
    pdf.set_font("Helvetica", size=12)
    for paragraph in content:
        pdf.multi_cell(0, 10, sanitize_text(paragraph))
        pdf.ln(5)
        
    # Output bytes
    pdf_bytes = pdf.output()
    if isinstance(pdf_bytes, str):
        return pdf_bytes.encode('latin1')
    return bytes(pdf_bytes)


# =====================================================================
# PREMIUM HOD PDF REPORT GENERATOR & HELPERS
# =====================================================================

class DepartmentReportPDF(FPDF):
    """
    Custom FPDF subclass implementing headers and footers on every page.
    """
    def __init__(self, title: str, generated_by: str):
        super().__init__()
        self.report_title = sanitize_text(title)
        self.generated_by = sanitize_text(generated_by)
        self.generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.set_title(self.report_title)

    def header(self):
        # Header title
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(100, 110, 120)
        self.cell(0, 5, sanitize_text("Department of Artificial Intelligence & Machine Learning"), new_x="LMARGIN", new_y="NEXT", align="L")
        
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(15, 23, 42)
        self.cell(0, 7, sanitize_text(self.report_title), new_x="LMARGIN", new_y="NEXT", align="L")
        
        self.set_font("Helvetica", "I", 9)
        self.set_text_color(148, 163, 184)
        self.cell(0, 5, sanitize_text(f"Generated: {self.generated_at} | By: {self.generated_by}"), new_x="LMARGIN", new_y="NEXT", align="L")
        self.ln(3)
        
        # Separator line
        self.set_draw_color(203, 213, 225)
        self.set_line_width(0.5)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(148, 163, 184)
        self.cell(0, 10, sanitize_text(f"Confidential — AIML Dept | Page {self.page_no()}/{{nb}}"), align="C")

    # Overridden methods to automatically sanitize text inputs
    def cell(self, w, h=0, txt="", *args, **kwargs):
        if "text" in kwargs:
            kwargs["text"] = sanitize_text(kwargs["text"])
        else:
            txt = sanitize_text(txt)
        return super().cell(w, h, txt, *args, **kwargs)

    def multi_cell(self, w, h=0, txt="", *args, **kwargs):
        if "text" in kwargs:
            kwargs["text"] = sanitize_text(kwargs["text"])
        else:
            txt = sanitize_text(txt)
        return super().multi_cell(w, h, txt, *args, **kwargs)

    def set_title(self, title: str, *args, **kwargs):
        return super().set_title(sanitize_text(title), *args, **kwargs)

def render_pdf_table(pdf: FPDF, table_rows: list[list[str]]):
    """
    Draws a standard bordered table grid from parsed markdown columns.
    """
    if not table_rows:
        return
    cols_count = max(len(row) for row in table_rows)
    page_width = 190.0 # 210mm total - 20mm margins
    col_width = page_width / cols_count
    
    for row_idx, row in enumerate(table_rows):
        if row_idx == 0:
            pdf.set_font("Helvetica", style="B", size=9)
            pdf.set_fill_color(241, 245, 249) # Light slate gray background for headers
        else:
            pdf.set_font("Helvetica", style="", size=9)
            pdf.set_fill_color(255, 255, 255)
            
        max_h = 7
        for col_idx in range(cols_count):
            val = row[col_idx].strip() if col_idx < len(row) else ""
            val = val.replace("**", "").replace("__", "")
            val = sanitize_text(val)
            pdf.cell(col_width, max_h, val, border=1, align="C", fill=True)
        pdf.ln(max_h)
    pdf.ln(4)

def write_markdown_line(pdf: FPDF, line_text: str, default_size: int = 10):
    """
    Parses headers, bold text segments, and lists for PDF rendering.
    """
    text = line_text.strip()
    if not text:
        return

    # Check for Headings
    if text.startswith("#"):
        hashes = len(text) - len(text.lstrip("#"))
        heading_text = text.lstrip("#").strip()
        size_boost = 4 if hashes == 1 else (3 if hashes == 2 else 2)
        
        pdf.set_font("Helvetica", style="B", size=default_size + size_boost)
        pdf.set_text_color(15, 23, 42)
        pdf.multi_cell(0, 8, sanitize_text(heading_text))
        pdf.ln(2)
        return

    # Check for Bullet List Item
    is_bullet = False
    if text.startswith("- ") or text.startswith("* "):
        is_bullet = True
        text = text[2:].strip()

    pdf.set_font("Helvetica", style="", size=default_size)
    pdf.set_text_color(51, 65, 85) # Slate-700 body text

    if is_bullet:
        pdf.set_x(15)
        pdf.write(5, "*  ")

    # Inline Bold Parser (**text**)
    parts = text.split("**")
    for idx, part in enumerate(parts):
        # Odd indices represent bold segments
        if idx % 2 == 1:
            pdf.set_font("Helvetica", style="B", size=default_size)
        else:
            pdf.set_font("Helvetica", style="", size=default_size)
        
        cleaned_part = sanitize_text(part)
        pdf.write(5, cleaned_part)
        
    pdf.ln(6)

def generate_pdf_report(title: str, content_markdown: str, metadata: dict) -> bytes:
    """
    Create FPDF object, set margins, parse markdown and return PDF bytes.
    """
    generated_by = metadata.get("generated_by", "Head of Department")
    
    # Pre-sanitize high-level input parameters
    title = sanitize_text(title)
    content_markdown = sanitize_text(content_markdown)
    generated_by = sanitize_text(generated_by)
    
    pdf = DepartmentReportPDF(title, generated_by)
    pdf.alias_nb_pages()
    pdf.set_margins(10, 10, 10)
    pdf.add_page()
    
    lines = content_markdown.split("\n")
    table_rows = []
    in_table = False
    
    for line in lines:
        line_stripped = line.strip()
        if line_stripped.startswith("|") and line_stripped.endswith("|"):
            # Check for header/body separator
            if all(c in "-:| " for c in line_stripped.replace("|", "")):
                continue
            cols = [col.strip() for col in line_stripped.split("|")[1:-1]]
            table_rows.append(cols)
            in_table = True
        else:
            if in_table:
                render_pdf_table(pdf, table_rows)
                table_rows = []
                in_table = False
            
            if line_stripped:
                write_markdown_line(pdf, line, default_size=10)
            else:
                pdf.ln(3)
                
    # Flush remaining table rows if any
    if in_table and table_rows:
        render_pdf_table(pdf, table_rows)
        
    pdf_bytes = pdf.output()
    if isinstance(pdf_bytes, str):
        return pdf_bytes.encode('latin1')
    return bytes(pdf_bytes)


# =====================================================================
# PREMIUM EXCEL REPORT GENERATORS
# =====================================================================

def generate_excel_report(records_dict: dict, sheet_name: str = None) -> bytes:
    """
    Generates a premium styled Excel report with frozen panes, merged headers,
    zebra striping and category summaries.
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 1. Summary Sheet
    summary_ws = wb.create_sheet(title="Summary")
    
    # Styling variables
    blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark blue
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Slate light gray
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Cool white-gray
    
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Title Row 1
    summary_ws.merge_cells("A1:C1")
    summary_ws["A1"] = "Department of Artificial Intelligence & Machine Learning"
    summary_ws["A1"].font = white_font
    summary_ws["A1"].fill = blue_fill
    summary_ws["A1"].alignment = Alignment(horizontal="center")
    
    # Row 2
    summary_ws["A2"] = "Academic Year Activity Report Summary"
    summary_ws["A2"].font = bold_font
    summary_ws["C2"] = f"Export Date: {datetime.now().strftime('%Y-%m-%d')}"
    summary_ws["C2"].font = bold_font
    
    # Row 3 Header
    summary_ws["A3"] = "Department Category"
    summary_ws["A3"].font = bold_font
    summary_ws["A3"].fill = header_fill
    summary_ws["A3"].border = thin_border
    
    summary_ws["B3"] = "Total Records"
    summary_ws["B3"].font = bold_font
    summary_ws["B3"].fill = header_fill
    summary_ws["B3"].border = thin_border
    
    # Fill Summary count data
    row_num = 4
    for key, val_list in records_dict.items():
        summary_ws.cell(row=row_num, column=1, value=str(key)).font = normal_font
        summary_ws.cell(row=row_num, column=1).border = thin_border
        summary_ws.cell(row=row_num, column=2, value=len(val_list)).font = normal_font
        summary_ws.cell(row=row_num, column=2).border = thin_border
        
        if row_num % 2 == 1:
            summary_ws.cell(row=row_num, column=1).fill = zebra_fill
            summary_ws.cell(row=row_num, column=2).fill = zebra_fill
        row_num += 1
        
    # Auto column width for Summary
    for col in summary_ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        summary_ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    summary_ws.freeze_panes = "A4"
    
    # 2. Individual sheets
    for cat_name, recs in records_dict.items():
        ws = wb.create_sheet(title=str(cat_name)[:30])
        
        if not recs:
            headers = ["Message"]
            data_rows = [["No submissions recorded"]]
        else:
            all_keys = []
            for r in recs:
                for k in r.keys():
                    if k not in all_keys and k != "doc_id":
                        all_keys.append(k)
            headers = all_keys
            data_rows = []
            for r in recs:
                row_vals = []
                for h in headers:
                    row_vals.append(r.get(h, ""))
                data_rows.append(row_vals)
                
        last_col_letter = get_column_letter(max(len(headers), 3))
        
        # Title row 1
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = "Department of Artificial Intelligence & Machine Learning"
        ws["A1"].font = white_font
        ws["A1"].fill = blue_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Row 2
        ws["A2"] = f"Collection: {cat_name}"
        ws["A2"].font = bold_font
        ws["B2"] = f"Export Date: {datetime.now().strftime('%Y-%m-%d')}"
        ws["B2"].font = bold_font
        
        # Row 3 Header
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=str(h).replace("_", " ").title())
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            
        # Data Rows
        for r_idx, row in enumerate(data_rows, 4):
            for c_idx, val in enumerate(row, 1):
                if isinstance(val, datetime):
                    val_str = val.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    val_str = str(val)
                cell = ws.cell(row=r_idx, column=c_idx, value=val_str)
                cell.font = normal_font
                cell.border = thin_border
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
        # Freeze top 3 rows
        ws.freeze_panes = "A4"
        
        # Auto column width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row == 1:
                    continue
                max_len = max(max_len, len(str(cell.value or '')))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    excel_out = io.BytesIO()
    wb.save(excel_out)
    return excel_out.getvalue()

def export_my_records_excel(records: dict, user_name: str, user_id: str) -> bytes:
    """
    Creates a personal Excel file for a faculty/student with their records.
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Ensure records is a dictionary
    if isinstance(records, list):
        records_dict = {"My Submissions": records}
    else:
        records_dict = records
        
    for sheet_name, rec_list in records_dict.items():
        ws = wb.create_sheet(title=str(sheet_name)[:30])
        
        if not rec_list:
            headers = ["Message"]
            data_rows = [["No personal records found"]]
        else:
            all_keys = []
            for r in rec_list:
                for k in r.keys():
                    if k not in all_keys and k != "doc_id":
                        all_keys.append(k)
            headers = all_keys
            data_rows = []
            for r in rec_list:
                row_vals = []
                for h in headers:
                    row_vals.append(r.get(h, ""))
                data_rows.append(row_vals)
                
        last_col_letter = get_column_letter(max(len(headers), 3))
        
        # Title merged row 1
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = "Department of Artificial Intelligence & Machine Learning"
        ws["A1"].font = white_font
        ws["A1"].fill = blue_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Row 2
        ws["A2"] = f"Personal Export: {user_name} ({user_id})"
        ws["A2"].font = bold_font
        ws["B2"] = f"Export Date: {datetime.now().strftime('%Y-%m-%d')}"
        ws["B2"].font = bold_font
        
        # Row 3 Header
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=str(h).replace("_", " ").title())
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            
        # Data
        for r_idx, row in enumerate(data_rows, 4):
            for c_idx, val in enumerate(row, 1):
                if isinstance(val, datetime):
                    val_str = val.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    val_str = str(val)
                cell = ws.cell(row=r_idx, column=c_idx, value=val_str)
                cell.font = normal_font
                cell.border = thin_border
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
        # Freeze top 3 rows
        ws.freeze_panes = "A4"
        
        # Auto column width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row == 1:
                    continue
                max_len = max(max_len, len(str(cell.value or '')))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    excel_out = io.BytesIO()
    wb.save(excel_out)
    return excel_out.getvalue()
